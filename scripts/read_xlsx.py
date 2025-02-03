import os
import openpyxl
from datetime import datetime

INPUT_DIR = "input"
OUTPUT_DIR = "output"

os.makedirs(OUTPUT_DIR, exist_ok=True)

def extract_question_and_options(desc):
    option_formats = ["a.", "b.", "c.", "d.", "(a)", "(b)", "(c)", "(d)","a)", "b)", "c)", "d)","A)","A.",
                      "1.", "2.", "3.", "4.", "(1)", "(2)", "(3)", "(4)"]
    options = []
    question = None

    parts = desc.split()
    current_option = None

    for part in parts:
        if any(part.startswith(fmt) for fmt in option_formats):
            if current_option is not None:
                options.append(" ".join(current_option).strip())
            current_option = [part]
        elif current_option is not None:
            current_option.append(part)
        else:
            question = part if question is None else f"{question} {part}"

    if current_option is not None:
        options.append(" ".join(current_option).strip())

    if len(options) != 4:
        raise ValueError("Could not find all four options in the description.")

    return question.strip(), options[0], options[1], options[2], options[3]

for file_name in os.listdir(INPUT_DIR):
    if file_name.endswith(".xlsx"):
        file_path = os.path.join(INPUT_DIR, file_name)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        processed_data = []
        skipped_data = []
        processed_count = 0
        skipped_count = 0

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            qno, desc, level, code, subject = row

            try:
                if isinstance(qno, str) and qno.startswith("="):
                    skipped_data.append((qno, desc))
                    skipped_count += 1
                    continue

                if not desc:
                    raise ValueError("Description is empty or missing.")

                question, option_a, option_b, option_c, option_d = extract_question_and_options(desc)

                processed_data.append((qno, question, option_a, option_b, option_c, option_d, level, code, subject))
                processed_count += 1

            except Exception as e:
                skipped_data.append((qno, desc))
                skipped_count += 1

        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Processed Data"

        output_sheet.append(["Qno", "Question", "Option A", "Option B", "Option C", "Option D", "Level", "Code", "Subject"])
        for row in processed_data:
            output_sheet.append(row)

        if skipped_data:
            skipped_sheet = output_workbook.create_sheet(title="Skipped Data")
            skipped_sheet.append(["Qno", "Description"])
            for row in skipped_data:
                skipped_sheet.append(row)

        date_suffix = datetime.now().strftime("%Y-%m-%d")
        output_file_name = f"{os.path.splitext(file_name)[0]}_processed_{date_suffix}.xlsx"
        output_file_path = os.path.join(OUTPUT_DIR, output_file_name)
        output_workbook.save(output_file_path)

        print(f"Processed data saved to: {output_file_path}")
        print(f"Processed rows: {processed_count}")
        print(f"Skipped rows: {skipped_count}")
